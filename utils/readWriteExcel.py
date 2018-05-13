# encoding:utf-8
from openpyxl import *
import os.path
from PyQt5.QtCore import *


class ReadWriteExcel(QObject):
    trigger = pyqtSignal(str, bool)

    def __init__(self):
        super(ReadWriteExcel, self).__init__()

    def writeExcel(self, fileName, uList, uTokenList, firstList, firstListToken, sheetName, sheetNameToken, num, sheetNameDiff, firstdiff, yesterday):
        if not os.path.isfile('{}.xlsx'.format(fileName)):
            wb = Workbook()
            wsDayInfo = wb.active
            ws2 = wb.create_sheet(sheetNameToken, 0) #token列表，最前插入
            wsDayInfo.title = sheetName #每日情况
            wsdiff = wb.create_sheet(sheetNameDiff) #每天变化表，在最后插入

            # 每天变化表
            wsdiff.cell(row=1, column=1, value=format(firstdiff[0]))
            wsdiff.cell(row=1, column=2, value=format(sheetNameToken + firstdiff[1]))

            # 每日情况
            for i in range(1, len(firstList) + 1):
                wsDayInfo.cell(row=1, column=i, value=format(firstList[i - 1]))
            # 每日情况
            for i in range(1, len(uTokenList) + 1):
                wsDayInfo.cell(row=2, column=i, value=format(uTokenList[i - 1]))

            # list表
            for i in range(1, len(firstListToken) + 1):
                ws2.cell(row=1, column=i, value=format(firstListToken[i - 1]))
            # list表
            number = num + 2
            #前500总币数
            total = 0
            for i in range(2, number):
                j = i - 2
                u = uList[j]
                for k in range(1, len(u) + 1):
                    ws2.cell(row=i, column=k, value=format(u[k - 1]))
                # 每天变化表
                if u[2] is None:
                    total = total + 0
                else:
                    total = total + eval(u[2])
                wsdiff.cell(row=i, column=1, value=format(u[1]))
                wsdiff.cell(row=i, column=2, value=format(u[2]))
            wsdiff.cell(row=number, column=1, value=format("总数"))
            wsdiff.cell(row=number, column=2, value=format(total))
            # self.creatediffsheet(wsdiff, firstdiff, sheetNameToken)
            # self.createdaysheet(firstList, wsDayInfo, uTokenList)
            # self.createtokensheet(ws2, firstListToken, num, uList, wsdiff)

            try:
                wb.save('{}.xlsx'.format(fileName))
                self.trigger.emit(fileName, True)
            except:
                self.trigger.emit(fileName, False)
        else:
            wb = load_workbook('{}.xlsx'.format(fileName))
            sheets = wb.sheetnames
            sheetDayInfo = wb[sheetName]
            yesterdaylist = [] #昨天的数据
            if sheetNameDiff not in sheets:
                sheetdiff = wb.create_sheet(sheetNameDiff)
                # 每天变化表
                sheetdiff.cell(row=1, column=1, value=format(firstdiff[0]))
                sheetdiff.cell(row=1, column=2, value=format(str(yesterday) + str(firstdiff[1])))

                #获取昨天数据
                if yesterday in sheets:
                    yesterdayInfo = wb[yesterday]  # 读取数据
                    for row in yesterdayInfo.rows:
                        # dist[row[1].value] = row[2].value

                        if row[2].value is None:
                            value = '0'
                        else:
                            value = row[2].value
                        yesterdaylist.append([row[1].value, value])
                else:

                    for u in uList:
                        yesterdaylist.append([u[1], u[2]])

                total = 0
                for i in range(1, len(yesterdaylist)):
                    # 每天变化表
                    sheetdiff.cell(row=i+1, column=1, value=format(yesterdaylist[i][0]))
                    sheetdiff.cell(row=i+1, column=2, value=format(yesterdaylist[i][1]))
                    if yesterdaylist[i][1] is not None:
                        total = total + eval(yesterdaylist[i][1])

                sheetdiff.cell(row=len(yesterdaylist)+1, column=1, value=format("总数"))
                sheetdiff.cell(row=len(yesterdaylist)+1, column=2, value=format(total))

            else:
                sheetdiff = wb[sheetNameDiff]
                # 读取昨天数据
                if sheetName in sheets:
                    maxcolum = sheetdiff.max_column
                    for row in sheetdiff.rows:
                        # dist[row[1].value] = row[2].value
                        #value = '0'
                        if maxcolum > 2:
                            if row[maxcolum - 2].value is None:
                                value = '0'
                            else:
                                value = row[maxcolum - 2].value
                        else:
                            if row[maxcolum - 1].value is None:
                                value = '0'
                            else:
                                value = row[maxcolum - 1].value
                        yesterdaylist.append([row[0].value, value])
                    yesterdaylist.pop() #删除最后一个元素{总数}

            diffmaxcolumn = sheetdiff.max_column+1

            #获取今天列表字典
            todydist = {}
            for u in uList:
                todydist[u[1]] = u[2]
            # print(yesterdaylist)
            # print("#"*150)
            # print(todydist)
            #获取变化数
            diffdata = []
            self.getdiffdate(yesterdaylist, todydist, diffdata)
            #print(yesterdaylist)
            #写入变化数
            sheetdiff.cell(row=1, column=diffmaxcolumn, value=format(str(sheetNameToken) + str(firstdiff[1])))
            sheetdiff.cell(row=1, column=diffmaxcolumn+1, value=format(firstdiff[2]))
            for i in range(len(diffdata)):
                sheetdiff.cell(row=i+2, column=diffmaxcolumn, value=format(diffdata[i][0]))
                sheetdiff.cell(row=i+2, column=diffmaxcolumn+1, value=format(diffdata[i][1]))


            maxRow = sheetDayInfo.max_row + 1
            for i in range(1, len(uTokenList) + 1):
                sheetDayInfo.cell(row=maxRow, column=i, value=format(uTokenList[i - 1]))
            if sheetNameToken not in sheets:
                wsNew = wb.create_sheet(sheetNameToken, 0)
                # list表
                for i in range(1, len(firstListToken) + 1):
                    wsNew.cell(row=1, column=i, value=format(firstListToken[i - 1]))
                number = num + 2

                for i in range(2, number):
                    j = i - 2
                    u = uList[j]
                    for k in range(1, len(u) + 1):
                        wsNew.cell(row=i, column=k, value=format(u[k - 1]))
            else:
                print("已存在sheet")
                sheetTokenList = wb[sheetNameToken]
                number = num + 2

                for i in range(2, number):
                    j = i - 2
                    u = uList[j]
                    for k in range(1, len(u) + 1):
                        sheetTokenList.cell(row=i, column=k, value=format(u[k - 1]))

            try:
                wb.save('{}.xlsx'.format(fileName))
                self.trigger.emit(fileName, True)
            except:
                self.trigger.emit(fileName, False)


    def getYesterdayData(self, filename, yesterday, yesterdaylist):
        #dist = {}
        #yesterdaylist = []
        if os.path.isfile('{}.xlsx'.format(filename)):
            sheetName = yesterday
            wb = load_workbook('{}.xlsx'.format(filename))
            sheets = wb.sheetnames
            if sheetName in sheets:
                yesterdayInfo = wb[yesterday]  # 读取数据
                for row in yesterdayInfo.rows:
                    #dist[row[1].value] = row[2].value
                    yesterdaylist.append([row[1].value, row[2].value])
                wb.close()
                #return True, dist, "读取昨天数据成功。"
                return True, "读取昨天数据成功。"
            else:
                wb.close()
                #return False, dist, "昨天的数据不存在。"
                return False, "昨天的数据不存在。"
        else:
            #return False, dist, "文件不存在。".format(filename)
            return False, "文件不存在。".format(filename)

    def getdiffdate(self, yesterdaylist, todydist, diffdata):
        newAddress = []
        total = 0
        totaldiff = 0
        for i in range(1, len(yesterdaylist)):
            # if todydist.get(yesterdaylist[i][0]) != None:
            #     diffvalue = int(round(eval(todydist.get(yesterdaylist[i][0])) - eval(yesterdaylist[i][1])))
            #     diffdata.append([todydist[yesterdaylist[i][0]], diffvalue])
            #     total = total + eval(todydist.get(yesterdaylist[i][0]))
            #     totaldiff = totaldiff + diffvalue
            # else:
            #     diffdata.append(['0', '0'])
            #     newAddress.append([i, yesterdaylist[i][0], yesterdaylist[i][1]])
            diffvalue = int(round(eval(todydist.get(yesterdaylist[i][0], '0')) - eval(yesterdaylist[i][1])))
            diffdata.append([todydist.get(yesterdaylist[i][0], '0'), diffvalue])
            total = total + eval(todydist.get(yesterdaylist[i][0], '0'))
            totaldiff = totaldiff + diffvalue
        diffdata.append([total, totaldiff])
        print(str(total) + ", " + str(totaldiff))





    # def creatediffsheet(self,wsdiff,firstdiff,sheetNameToken):
    #     # 每天变化表
    #     wsdiff.cell(row=1, column=1, value=format(firstdiff[0]))
    #     wsdiff.cell(row=1, column=2, value=format(sheetNameToken + firstdiff[1]))
    #
    # def createtokensheet(self, ws2, firstListToken, num, uList, wsdiff):
    #     # list表
    #     for i in range(1, len(firstListToken) + 1):
    #         ws2.cell(row=1, column=i, value=format(firstListToken[i - 1]))
    #     # list表
    #     number = num + 2
    #     for i in range(2, number):
    #         j = i - 2
    #         u = uList[j]
    #         for k in range(1, len(u) + 1):
    #             ws2.cell(row=i, column=k, value=format(u[k - 1]))
    #         # 每天变化表
    #         wsdiff.cell(row=i, column=1, value=format(u[1]))
    #         wsdiff.cell(row=i, column=2, value=format(u[2]))
    #
    # def createdaysheet(self, firstList, wsDayInfo, uTokenList):
    #     # 每日情况
    #     for i in range(1, len(firstList) + 1):
    #         wsDayInfo.cell(row=1, column=i, value=format(firstList[i - 1]))
    #     # 每日情况
    #     for i in range(1, len(uTokenList) + 1):
    #         wsDayInfo.cell(row=2, column=i, value=format(uTokenList[i - 1]))

